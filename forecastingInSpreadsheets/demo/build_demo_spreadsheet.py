"""
Build the Demo 1.2 spreadsheet (bakery_sales_demo.xlsx) with:
- Raw data
- Naive forecast column
- 7-day and 30-day moving average columns
- TREND forecast column
- 14-day forecast extension rows
- Chart: last 90 days of actuals + 14 days of all forecasts
"""

import csv
import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
data_path = Path(__file__).parent.parent / "data" / "bakery_sales.csv"
rows = []
with open(data_path) as f:
    reader = csv.DictReader(f)
    for r in reader:
        rows.append((datetime.date.fromisoformat(r["date"]), int(r["loaves_sold"])))

N = len(rows)  # 731
FORECAST_DAYS = 14

# ---------------------------------------------------------------------------
# Create workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bakery Forecast"

# -- Header row --
headers = ["Date", "Loaves Sold", "Naive Forecast", "7-Day MA", "30-Day MA", "Trend Forecast"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font

# -- Data rows (rows 2 through N+1) --
for i, (date, sales) in enumerate(rows):
    row_num = i + 2  # row 2 = first data row
    ws.cell(row=row_num, column=1, value=date).number_format = "YYYY-MM-DD"
    ws.cell(row=row_num, column=2, value=sales)

last_data_row = N + 1  # row 732

# -- Forecast extension rows (rows N+2 through N+1+FORECAST_DAYS) --
last_date = rows[-1][0]
for d in range(1, FORECAST_DAYS + 1):
    row_num = last_data_row + d
    future_date = last_date + datetime.timedelta(days=d)
    ws.cell(row=row_num, column=1, value=future_date).number_format = "YYYY-MM-DD"

last_forecast_row = last_data_row + FORECAST_DAYS  # row 746

# ---------------------------------------------------------------------------
# Formulas — using Google-Sheets-compatible syntax where possible
# ---------------------------------------------------------------------------

# Naive forecast: last observed value repeated for all 14 forecast days
for row_num in range(last_data_row + 1, last_forecast_row + 1):
    ws.cell(row=row_num, column=3, value=f"=B{last_data_row}")

# 7-day moving average: average of last 7 actual values, repeated
for row_num in range(last_data_row + 1, last_forecast_row + 1):
    ws.cell(row=row_num, column=4, value=f"=AVERAGE(B{last_data_row-6}:B{last_data_row})")

# 30-day moving average: average of last 30 actual values, repeated
for row_num in range(last_data_row + 1, last_forecast_row + 1):
    ws.cell(row=row_num, column=5, value=f"=AVERAGE(B{last_data_row-29}:B{last_data_row})")

# Trend forecast: TREND over all data, projected to each forecast date
# In Excel/Sheets: =TREND(known_y, known_x_as_numbers, new_x)
# We use column A (dates) as x. Excel treats dates as serial numbers, so TREND works.
for row_num in range(last_data_row + 1, last_forecast_row + 1):
    ws.cell(
        row=row_num,
        column=6,
        value=f"=TREND(B$2:B${last_data_row},A$2:A${last_data_row},A{row_num})",
    )

# Also fill trend for the actual data period (so the trend line shows on the chart)
for row_num in range(2, last_data_row + 1):
    ws.cell(
        row=row_num,
        column=6,
        value=f"=TREND(B$2:B${last_data_row},A$2:A${last_data_row},A{row_num})",
    )

# ---------------------------------------------------------------------------
# Chart: last 90 days of actuals + 14 days of forecasts
# ---------------------------------------------------------------------------
chart = LineChart()
chart.title = "Bakery Sales: Actuals vs. Forecasts"
chart.y_axis.title = "Loaves Sold"
chart.x_axis.title = "Date"
chart.style = 10
chart.width = 28
chart.height = 14

# Chart range: last 90 data rows + 14 forecast rows = 104 rows
chart_start_row = last_data_row - 89  # 90 days back from last data row
chart_end_row = last_forecast_row

# X axis (dates)
dates_ref = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)

# Actuals (column B)
actuals = Reference(ws, min_col=2, min_row=chart_start_row - 1, max_row=chart_end_row)
chart.add_data(actuals, titles_from_data=True, from_rows=False)

# Naive (column C)
naive = Reference(ws, min_col=3, min_row=chart_start_row - 1, max_row=chart_end_row)
chart.add_data(naive, titles_from_data=True)

# 7-day MA (column D)
ma7 = Reference(ws, min_col=4, min_row=chart_start_row - 1, max_row=chart_end_row)
chart.add_data(ma7, titles_from_data=True)

# 30-day MA (column E)
ma30 = Reference(ws, min_col=5, min_row=chart_start_row - 1, max_row=chart_end_row)
chart.add_data(ma30, titles_from_data=True)

# Trend (column F)
trend = Reference(ws, min_col=6, min_row=chart_start_row - 1, max_row=chart_end_row)
chart.add_data(trend, titles_from_data=True)

chart.set_categories(dates_ref)

# Style the series
colors = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "70AD47"]
for idx, color in enumerate(colors):
    s = chart.series[idx]
    s.graphicalProperties.line.width = 20000 if idx == 0 else 15000

ws.add_chart(chart, "H2")

# ---------------------------------------------------------------------------
# Column widths
# ---------------------------------------------------------------------------
ws.column_dimensions["A"].width = 14
for col in range(2, 7):
    ws.column_dimensions[get_column_letter(col)].width = 16

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
out_path = Path(__file__).parent / "bakery_sales_demo.xlsx"
wb.save(out_path)
print(f"Created: {out_path}")
print(f"  Data rows: {N}")
print(f"  Forecast rows: {FORECAST_DAYS}")
print(f"  Chart covers rows {chart_start_row}-{chart_end_row} (last 90 days + 14 forecast days)")
print()
print("To use:")
print("  1. Upload bakery_sales_demo.xlsx to Google Sheets")
print("  2. The formulas will auto-convert")
print("  3. The chart will render in Sheets")
