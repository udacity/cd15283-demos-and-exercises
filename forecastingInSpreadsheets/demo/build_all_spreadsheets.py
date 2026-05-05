"""
Build all three Exercise 1.2 spreadsheets:
  - demo/bakery_sales_demo.xlsx    (fully built, for screen recording)
  - starter/bakery_sales_starter.xlsx  (data + empty columns, for students)
  - solution/bakery_sales_solution.xlsx (fully built, for answer key)
"""

import csv
import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
data_path = Path(__file__).parent.parent / "data" / "bakery_sales.csv"
rows = []
with open(data_path) as f:
    reader = csv.DictReader(f)
    for r in reader:
        rows.append((datetime.date.fromisoformat(r["date"]), int(r["loaves_sold"])))

N = len(rows)  # 731
FORECAST_DAYS = 14


def build_spreadsheet(out_path: Path, include_formulas: bool, include_chart: bool):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bakery Forecast"

    # -- Headers --
    headers = ["Date", "Loaves Sold", "Naive Forecast", "7-Day MA", "30-Day MA", "Trend Forecast"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # -- Data rows --
    for i, (date, sales) in enumerate(rows):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=date).number_format = "YYYY-MM-DD"
        ws.cell(row=row_num, column=2, value=sales)

    last_data_row = N + 1  # 732

    # -- Forecast extension rows --
    last_date = rows[-1][0]
    for d in range(1, FORECAST_DAYS + 1):
        row_num = last_data_row + d
        future_date = last_date + datetime.timedelta(days=d)
        ws.cell(row=row_num, column=1, value=future_date).number_format = "YYYY-MM-DD"

    last_forecast_row = last_data_row + FORECAST_DAYS  # 746

    if include_formulas:
        # Naive forecast
        for row_num in range(last_data_row + 1, last_forecast_row + 1):
            ws.cell(row=row_num, column=3, value=f"=B{last_data_row}")

        # 7-day MA
        for row_num in range(last_data_row + 1, last_forecast_row + 1):
            ws.cell(row=row_num, column=4, value=f"=AVERAGE(B{last_data_row-6}:B{last_data_row})")

        # 30-day MA
        for row_num in range(last_data_row + 1, last_forecast_row + 1):
            ws.cell(row=row_num, column=5, value=f"=AVERAGE(B{last_data_row-29}:B{last_data_row})")

        # Trend — for the full data range + forecast
        for row_num in range(2, last_forecast_row + 1):
            ws.cell(
                row=row_num,
                column=6,
                value=f"=TREND(B$2:B${last_data_row},A$2:A${last_data_row},A{row_num})",
            )

    if include_chart:
        chart = LineChart()
        chart.title = "Bakery Sales: Actuals vs. Forecasts"
        chart.y_axis.title = "Loaves Sold"
        chart.x_axis.title = "Date"
        chart.style = 10
        chart.width = 28
        chart.height = 14

        chart_start_row = last_data_row - 89
        chart_end_row = last_forecast_row

        dates_ref = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)

        for col in range(2, 7):
            data_ref = Reference(ws, min_col=col, min_row=chart_start_row - 1, max_row=chart_end_row)
            chart.add_data(data_ref, titles_from_data=True)

        chart.set_categories(dates_ref)
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.numFmt = "YYYY-MM-DD"
        chart.x_axis.majorUnit = 30
        ws.add_chart(chart, "H2")

    # Column widths
    ws.column_dimensions["A"].width = 14
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 16

    wb.save(out_path)
    label = "with formulas + chart" if include_formulas else "starter (no formulas)"
    print(f"  Created: {out_path}  [{label}]")


# ---------------------------------------------------------------------------
# Build all three
# ---------------------------------------------------------------------------
base = Path(__file__).parent.parent
print("Building Exercise 1.2 spreadsheets...\n")

build_spreadsheet(
    base / "demo" / "bakery_sales_demo.xlsx",
    include_formulas=True,
    include_chart=True,
)

build_spreadsheet(
    base / "starter" / "bakery_sales_starter.xlsx",
    include_formulas=False,
    include_chart=False,
)

build_spreadsheet(
    base / "solution" / "bakery_sales_solution.xlsx",
    include_formulas=True,
    include_chart=True,
)

print("\nDone! Upload to Google Sheets — formulas auto-convert.")
